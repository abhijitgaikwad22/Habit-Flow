from flask import Flask, render_template, request, redirect, session, jsonify, send_from_directory
import json, os
from datetime import datetime, date, timedelta

app = Flask(__name__)
app.secret_key = 'habitflow_secret_2024'

USERS_FILE   = 'users.json'
HABITS_FILE  = 'habits.json'
UPLOAD_FOLDER = 'uploads'

# Create uploads folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def load_json(file):
    if os.path.exists(file):
        with open(file, 'r') as f:
            return json.load(f)
    return {}

def save_json(file, data):
    with open(file, 'w') as f:
        json.dump(data, f, indent=2)

def today_str():
    return date.today().isoformat()

def get_user_habits(username):
    return load_json(HABITS_FILE).get(username, [])

def save_user_habits(username, user_habits):
    habits = load_json(HABITS_FILE)
    habits[username] = user_habits
    save_json(HABITS_FILE, habits)

def calc_streak(completions):
    if not completions: return 0
    streak, check = 0, date.today()
    while check.isoformat() in completions:
        streak += 1
        check -= timedelta(days=1)
    return streak

def get_weekly_data(habits):
    days = [(date.today() - timedelta(days=i)).isoformat() for i in range(6, -1, -1)]
    total = len(habits)
    result = []
    for d in days:
        done = sum(1 for h in habits if d in h.get('completions', []))
        pct  = int((done / total * 100)) if total else 0
        result.append({
            'date':  d,
            'label': datetime.fromisoformat(d).strftime('%a'),
            'done':  done,
            'total': total,
            'pct':   pct,
        })
    return result


@app.route('/register', methods=['GET', 'POST'])
def register():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '').strip()
        users    = load_json(USERS_FILE)
        if not username or not password:
            error = 'Please fill in all fields.'
        elif username in users:
            error = 'Username already taken.'
        elif len(password) < 4:
            error = 'Password must be at least 4 characters.'
        else:
            users[username] = {'password': password, 'joined': today_str()}
            save_json(USERS_FILE, users)
            session['user'] = username
            return redirect('/')
    return render_template('register.html', error=error)


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '').strip()
        users    = load_json(USERS_FILE)
        if username in users and users[username]['password'] == password:
            session['user'] = username
            return redirect('/')
        else:
            error = 'Invalid username or password.'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect('/login')


@app.route('/')
def index():
    if 'user' not in session:
        return redirect('/login')

    username   = session['user']
    habits     = get_user_habits(username)
    today      = today_str()
    category_f = request.args.get('category', 'All')
    search_q   = request.args.get('search', '').lower()

    for h in habits:
        h['done_today'] = today in h.get('completions', [])
        h['streak']     = calc_streak(h.get('completions', []))
        h['total_done'] = len(h.get('completions', []))
        goal            = h.get('goal_days', 0)
        h['goal_pct']   = min(int((h['total_done'] / goal * 100)), 100) if goal else 0

    filtered = habits
    if category_f != 'All':
        filtered = [h for h in filtered if h.get('category') == category_f]
    if search_q:
        filtered = [h for h in filtered if search_q in h['title'].lower()]

    total       = len(habits)
    done_today  = sum(1 for h in habits if h['done_today'])
    pct         = int((done_today / total * 100)) if total else 0
    best_streak = max((h['streak'] for h in habits), default=0)
    total_ever  = sum(h['total_done'] for h in habits)
    weekly_data = get_weekly_data(habits)
    week_json   = json.dumps(weekly_data)

    return render_template('index.html',
        username=username,
        habits=filtered,
        today=today,
        total=total,
        done_today=done_today,
        pct=pct,
        best_streak=best_streak,
        total_ever=total_ever,
        categories=['All', 'Health', 'Learning', 'Work', 'Personal', 'Fitness'],
        active_cat=category_f,
        search_q=search_q,
        weekly_data=weekly_data,
        week_json=week_json,
    )


@app.route('/add', methods=['POST'])
def add_habit():
    if 'user' not in session: return redirect('/login')
    username  = session['user']
    title     = request.form.get('title', '').strip()
    category  = request.form.get('category', 'Personal')
    freq      = request.form.get('frequency', 'Daily')
    note      = request.form.get('note', '').strip()
    goal_days = request.form.get('goal_days', '').strip()
    if title:
        habits = get_user_habits(username)
        habits.append({
            'id':          len(habits) + 1,
            'title':       title,
            'category':    category,
            'frequency':   freq,
            'note':        note,
            'goal_days':   int(goal_days) if goal_days.isdigit() else 0,
            'completions': [],
            'created':     today_str(),
        })
        save_user_habits(username, habits)
    return redirect('/')


@app.route('/toggle/<int:habit_id>')
def toggle(habit_id):
    if 'user' not in session: return redirect('/login')
    username = session['user']
    habits   = get_user_habits(username)
    today    = today_str()
    for h in habits:
        if h['id'] == habit_id:
            if today in h['completions']: h['completions'].remove(today)
            else: h['completions'].append(today)
            break
    save_user_habits(username, habits)
    return redirect('/')


@app.route('/delete/<int:habit_id>')
def delete_habit(habit_id):
    if 'user' not in session: return redirect('/login')
    username = session['user']
    habits   = [h for h in get_user_habits(username) if h['id'] != habit_id]
    save_user_habits(username, habits)
    return redirect('/')


@app.route('/profile')
def profile():
    if 'user' not in session: return redirect('/login')
    username = session['user']
    habits   = get_user_habits(username)
    users    = load_json(USERS_FILE)
    joined   = users.get(username, {}).get('joined', 'Unknown')
    photo    = users.get(username, {}).get('photo', None)
    today    = today_str()

    for h in habits:
        h['streak']     = calc_streak(h.get('completions', []))
        h['total_done'] = len(h.get('completions', []))
        h['done_today'] = today in h.get('completions', [])

    cat_breakdown = {}
    for h in habits:
        cat = h.get('category', 'Personal')
        if cat not in cat_breakdown:
            cat_breakdown[cat] = {'total': 0, 'done': 0, 'pct': 0}
        cat_breakdown[cat]['total'] += 1
        if h['done_today']:
            cat_breakdown[cat]['done'] += 1
    for cat in cat_breakdown:
        t = cat_breakdown[cat]['total']
        d = cat_breakdown[cat]['done']
        cat_breakdown[cat]['pct'] = int((d / t * 100)) if t else 0

    top_habits        = sorted(habits, key=lambda h: h['streak'], reverse=True)[:5]
    total_habits      = len(habits)
    total_completions = sum(h['total_done'] for h in habits)
    best_streak       = max((h['streak'] for h in habits), default=0)

    return render_template('profile.html',
        username=username,
        joined=joined,
        photo=photo,
        total_habits=total_habits,
        total_completions=total_completions,
        best_streak=best_streak,
        cat_breakdown=cat_breakdown,
        top_habits=top_habits,
    )


@app.route('/upload_photo', methods=['POST'])
def upload_photo():
    if 'user' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    username = session['user']
    if 'photo' not in request.files:
        return jsonify({'success': False, 'error': 'No file'})
    file = request.files['photo']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'})

    # Only allow images
    allowed = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in allowed:
        return jsonify({'success': False, 'error': 'Invalid file type'})

    # Save as username.ext
    filename = f"{username}.{ext}"
    file.save(os.path.join(UPLOAD_FOLDER, filename))

    # Save photo filename in users.json
    users = load_json(USERS_FILE)
    if username in users:
        users[username]['photo'] = filename
        save_json(USERS_FILE, users)

    return jsonify({'success': True, 'filename': filename})


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


# ── ACHIEVEMENTS ─────────────────────────────────────
ACHIEVEMENTS = [
    {'id':'first_habit',  'icon':'🌱', 'title':'First Step',    'desc':'Add your first habit',             'check': lambda h,s: len(h) >= 1},
    {'id':'streak_3',     'icon':'🔥', 'title':'On Fire',       'desc':'Reach a 3-day streak',             'check': lambda h,s: s >= 3},
    {'id':'streak_7',     'icon':'⚡', 'title':'Week Warrior',  'desc':'Reach a 7-day streak',             'check': lambda h,s: s >= 7},
    {'id':'streak_30',    'icon':'💎', 'title':'Diamond Habit', 'desc':'Reach a 30-day streak',            'check': lambda h,s: s >= 30},
    {'id':'five_habits',  'icon':'🎯', 'title':'Habit Builder', 'desc':'Track 5 or more habits',           'check': lambda h,s: len(h) >= 5},
    {'id':'total_50',     'icon':'💪', 'title':'Half Century',  'desc':'Complete 50 total check-ins',      'check': lambda h,s: sum(len(x.get('completions',[])) for x in h) >= 50},
    {'id':'total_100',    'icon':'🏆', 'title':'Century Club',  'desc':'Complete 100 total check-ins',     'check': lambda h,s: sum(len(x.get('completions',[])) for x in h) >= 100},
    {'id':'all_cats',     'icon':'🌈', 'title':'All Rounder',   'desc':'Have habits in all 5 categories',  'check': lambda h,s: len(set(x.get('category') for x in h)) >= 5},
]

@app.route('/achievements')
def achievements():
    if 'user' not in session: return redirect('/login')
    username    = session['user']
    habits      = get_user_habits(username)
    best_streak = max((calc_streak(h.get('completions',[])) for h in habits), default=0)
    result = []
    for a in ACHIEVEMENTS:
        unlocked = a['check'](habits, best_streak)
        result.append({'id':a['id'],'icon':a['icon'],'title':a['title'],'desc':a['desc'],'unlocked':unlocked})
    unlocked_count = sum(1 for a in result if a['unlocked'])
    return render_template('achievements.html',
        username=username, achievements=result,
        unlocked_count=unlocked_count, total_count=len(result))


# ── CALENDAR ─────────────────────────────────────────
@app.route('/calendar/<int:habit_id>')
def calendar(habit_id):
    if 'user' not in session: return redirect('/login')
    username = session['user']
    habits   = get_user_habits(username)
    habit    = next((h for h in habits if h['id'] == habit_id), None)
    if not habit: return redirect('/')
    habit['streak']     = calc_streak(habit.get('completions', []))
    habit['total_done'] = len(habit.get('completions', []))
    today_d  = date.today()
    start    = today_d - timedelta(weeks=12)
    cal_days = []
    cur = start
    while cur <= today_d:
        cal_days.append({
            'date':  cur.isoformat(),
            'day':   cur.day,
            'dow':   cur.weekday(),
            'done':  cur.isoformat() in habit.get('completions', []),
            'today': cur == today_d,
        })
        cur += timedelta(days=1)
    return render_template('calendar.html',
        username=username, habit=habit, cal_days=cal_days, today=today_str())


# ── EXPORT EXCEL ──────────────────────────────────────
@app.route('/export')
def export():
    if 'user' not in session: return redirect('/login')
    from flask import Response
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    username = session['user']
    habits   = get_user_habits(username)
    today    = today_str()

    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Summary'

    ws1.merge_cells('A1:H1')
    ws1['A1'] = f'HabitFlow Report — {username} — {today}'
    ws1['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws1['A1'].fill      = PatternFill('solid', fgColor='EA580C')
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 28

    headers = ['Habit','Category','Frequency','Streak','Total Done','Created','Note','Goal Days']
    for i, h in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', fgColor='1A1208')
        cell.alignment = Alignment(horizontal='center')

    for row, h in enumerate(habits, 3):
        data = [h.get('title',''), h.get('category',''), h.get('frequency',''),
                calc_streak(h.get('completions',[])), len(h.get('completions',[])),
                h.get('created',''), h.get('note',''), h.get('goal_days',0)]
        for col, val in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal='left', wrap_text=True)
            if row % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='1C1510')

    for col in ws1.columns:
        w = max((len(str(c.value)) for c in col if c.value), default=10)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = min(w+4, 40)

    ws2 = wb.create_sheet('Completion History')
    ws2['A1'] = 'Habit'
    ws2['B1'] = 'Date'
    ws2['A1'].font = ws2['B1'].font = Font(bold=True, color='FFFFFF')
    ws2['A1'].fill = ws2['B1'].fill = PatternFill('solid', fgColor='EA580C')
    row = 2
    for h in habits:
        for d in sorted(h.get('completions', [])):
            ws2.cell(row=row, column=1, value=h['title'])
            ws2.cell(row=row, column=2, value=d)
            row += 1
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=habitflow_{username}_{today}.xlsx'})


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)